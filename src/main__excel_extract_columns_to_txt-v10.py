# Lao dictionary
# ('ROMANIZED THAI', 'EASYTHAI', 'THAIPHON', 'THA', 'LAO1', 'LAO2', 'ENG', 'TYPE', 'USAGE', 'SCIENT', 'CLASSIF', 'NOTE')

# thai dictionary

# volubilis mundo (all 15 lang) - the file 'vol..database doesnt have the lao..)

# 0'THAIROM, 1'EASYTHAI, 2'THAIPHON, 3'THA (Thai), ![4 added: thai_pron] , 5'ENG (English), 6'FRA (French), 7'TYPE, 8'USAGE, 9'SCIENT/abbrev., 10'DOM, 11'CLASSIF, 12'SYN, 13'LEVEL, 14'NOTE, 15'SPA (Spanish), 16'ITA , 17'POR (Portuguese), 18'DEU (German), 19'NLD, 20'NOR (Norwegian), 21'TUR , 22'MSA , 23'IND (Indonesian), 24'FIL , 25'VIE (Vietnamese), 26'RUS1 , 27'RUS2, 28'LAO1 (Lao), 29'LAO2 (Lao), 30'KOR1 (Korean), 31'KOR2 (Korean),


import pdb
import os
import regex
from collections import OrderedDict




re_remove_brackets = OrderedDict(
    [
        (r"[…\.\[\]\(\)]", ""),  #
        (r"[\u0300\u0301\u0302\u030c]", ""),  #
        # (r'^\.\.\.',''),              #
        # (r'^[…\(\[\.]',''),              #
        # (r'[\)\]]$',''),              #
    ]
)

re_remove_starting_brackets = OrderedDict(
    [
        # (r'^\.\.\.',''),              #
        (r"\[", "("),  #
        (r"\]", ")"),  #
        (r"^\((.*)\)$", "\\1"),  #
        (r"([-_\\/¯]) ([a-zɔ])", r"\1\2"),  #
    ]
)

re_type_friendly_1 = OrderedDict(
    [
        (r"ø̅", "øø"),  #
        (r"ɔ̅", "ɔɔ"),  #
        (r"ē", "ee"),  #
        (r"ā", "aa"),  #
        (r"ī", "ii"),  #
        (r"ū", "uu"),  #
        (r"ō", "oo"),  #
        (r"", ""),  #
    ]
)


# à  _   \u0300
# á  ¯   \u0301
# â  \   \u0302
# ǎ  /   \u030c
# r'\n' = '\\n'
#  ' <some_char> '.encode('utf-8')
re_to_paiboon = OrderedDict(
    [
        (r"¯([bcdfghjklmnpqrstvwxyz]*)([aeiouɔ])([a-z]*)", r"-\1\2\u0301\3"),  #
        (r"\\([bcdfghjklmnpqrstvwxyz]*)([aeiouɔ])([a-z]*)", r"-\1\2\u0302\3"),  #
        (r"/([bcdfghjklmnpqrstvwxyz]*)([aeiouɔ])([a-z]*)", r"-\1\2\u030c\3"),  #
        (r"_([bcdfghjklmnpqrstvwxyz]*)([aeiouɔ])([a-z]*)", r"-\1\2\u0300\3"),  #
        (r"^-", ""),  #
        (r"([\(\[])-", r"\1"),  # (-maa -> (maa
    ]
)

re_spaces_workaround_dictbox = OrderedDict(
    [(
            r"[ ]-",
            r"<sp> </sp>",
    )]
    # dictbox ignores spaces if the string inside a tag is non latin
    # (eg diacrytics.. workaround: add a tag with a single whitespace..

)

re_type_friendly_2 = OrderedDict(
    [
        (r"[-_\\/¯]", " "),  # remove toneguide for tone search
        (r"^\s+", ""),  # remove leading whitespace
        (r"\s\s+", " "),  # rm double spaces
        (r"([tkp])h", "\\1"),  # simplify search for aspirated an un- letters
        (r"ñ", "n"),  # ñ as n because its not on the default keyboard
        (r"ɔ", "o"),  # same for ɔ
        (r"", ""),  #
    ]
)

re_pron = OrderedDict(
    [
        (r"ø", "ø̅"),  # continuous use of upper bar
        (r"ǿ", "ø"),  # same as above
        (r"ø̅", "ɔ̅"),  # optional: ɔ is more standart than ø..
    ]
)
# needs a separate var, because ø is already defined an woul only be overridden.. hard to find that one!
re_pron2 = OrderedDict(
    [
        (r"ø", "ɔ"),
        (r"", ""),
        (r"", ""),
    ]
)
re_default = OrderedDict(
    [
        (r"\t", "    "),  # no tabs allowed
        (r"^\s+", ""),  # remove leading whitespace
        (r"\s+$", ""),  # remove tailing whitespace
    ]
)

re_final_pron = OrderedDict(
    [
        (
            r"\\",
            "&#x5c;",
        ),  # substitute slashes with html notation &#x<hex-value>; eg \=5c (g8 in vim)
    ]
)

re_classifier = OrderedDict(
    [
        # (r'\s*(.*?)\s*\(([ก-๛]+).*\s*','\\2 (\\1)'),           #both thai & pron, but pron is broken.. no tones
        (r"\s*(.*?)\s*\(([ก-๛]+).*\s*", "\\2"),
        (r"", ""),
        (r"", ""),
    ]
)


def replace_multi(text, dic, debug=False):
    t = text
    for i, j in dic.items():
        #  p1 = regex.compile(i)
        # if(i==''): continue
        # if(debug): print(i+' -> '+t+' '+text)
        text = regex.sub(i, j, text)
        if debug:
            print(i + " ->" + t + " " + text)
    # breakpoint()
    return text


def format_tones(s, paiboon=False):
    s = replace_multi(s, re_default)
    s = replace_multi(s, re_remove_starting_brackets)
    s = replace_multi(s, re_pron)
    s = replace_multi(s, re_pron2)
    if paiboon:
        s = replace_multi(s, re_type_friendly_1)
        s = replace_multi(s, re_to_paiboon)
    return s


def format_def(s):
    s = replace_multi(s, re_default)
    return s


def format_final_pron(s, paiboon=False):
    if not paiboon:
        s = replace_multi(s, re_final_pron)
    return s


def format_classifier(s, paiboon=False):
    s = replace_multi(s, re_classifier)
    s = replace_multi(s, re_default)
    s = replace_multi(s, re_remove_starting_brackets)
    s = replace_multi(s, re_pron)
    s = replace_multi(s, re_pron2)
    if paiboon:
        s = replace_multi(s, re_type_friendly_1)
        s = replace_multi(s, re_to_paiboon)
    return s


def spaces_workaround_dictbox(s):
    s = replace_multi(s, re_spaces_workaround_dictbox)
    return s


def format_pron_search(s, paiboon=False):
    s = replace_multi(s, re_remove_brackets)
    if not paiboon:
        s = replace_multi(s, re_type_friendly_1)
    s = replace_multi(s, re_type_friendly_2)
    return s


def prepend_to_file(file='', lines=[]):
    tmp = ''
    with open(file, 'r') as f:
        tmp = f.read()
    with open(file, 'w') as f:
        f.write(('\n').join(lines)+'\n')
        f.write(tmp)


def append_to_file(file='', lines=[]):
    with open(file, 'a') as f:
        f.write(('\n').join(lines))


def make_dict_files():
    titleE = 'volubilis v074 (en-th)'
    titleTe = 'volubilis v074 (.th-en)'
    titleT = 'volubilis v074 (th-en)'
    description = 'description=Volubilis English-Thai dictionary by Belisan (Fr. Bastien)<br>พจนานุกรม วอลุบิลิส ภาษาอังกฤษ-ไทย<br>v. 21.3 (1.11.2022) - 103 000  entr.<br>(http://belisan-volubilis.blogspot.com)<br><br>ā = long vowel "a"<br>start pronounciation search: type . (dot) plus searchterm (eg. .maa -> dog, horse,come,..)'
    out_path = 'vol_mundo/volubilis_v074'
    os.system('mkdir -r ' + out_path)
    fi = 'vol_mundo/d_'
    fo = out_path + '/volubilis_'
    conv_to_ifo(file_in=fi+'en-th.txt', file_out=fo+'(en-th).ifo', name=titleE, description=description)
    conv_to_ifo(file_in=fi+'th(dot+pr)-en.txt', file_out=fo+'(th_pron-en).ifo', name=titleTe, description=description)
    conv_to_ifo(file_in=fi+'th-en.txt', file_out=fo+'(th-en).ifo', name=titleT, description=description)
    append_to_file(file=fo+'(en-th).ifo', lines=['src=th', 'dst=th'])
    append_to_file(file=fo+'(th-en).ifo', lines=['src=th', 'dst=th'])
    append_to_file(file=fo+'(th_pron-en).ifo', lines=['src=th', 'dst=th'])

    print('\n\n      --!-- files copied to + overridden in dictboxdata --!--\n')
    os.system("rm '/storage/emulated/0/dictboxdata/03 thai/volubilis_v074/'* ")
    os.system("cp -r '/storage/emulated/0/python/dict/" + out_path + "' '/storage/emulated/0/dictboxdata/03 thai/' ")




def conv_to_ifo(file_in, file_out, name, description, use_imported_lib=False, use_proot_distro=False):
    print('\n  --pyglossary (+dictzip): '+file_in)
    # set title and description (no commandline option in pyglossary for description)
    prepend_to_file(file_in, ['##title\t'+name, '##description\t'+description])

    # start proot-distro and execute pyglossary (superslow)
    i = os.getcwd()+'/'+file_in
    o = os.getcwd()+'/'+file_out

    # sort the file (faster finding?) and -v2 sets verbrosity to only errors and warnings
    os.system('pyglossary "'+i+'" "'+o+'" --sort -v1')


# working
def import_excel_file(xlsx_file, to_folder, columns=30, paiboon=True, debug=False):

    from openpyxl import load_workbook  # for excel xlsx files

    # read_only option reduces memory. openpyxl lib takes huge amounts in read+write mode..
    wb2 = load_workbook(xlsx_file, read_only=True)
    print(wb2.sheetnames)

    r = []

    folder = to_folder

    try:
        os.mkdir(folder)
    except Exception:
        print("folder is there")
    if debug:
        for i in range(0, columns):
            r.append(open(folder + "/r_" + str(i).zfill(2) + ".txt", "w"))
    f_th_en = open(folder + "/d_th-en.txt", "w")
    f_pr_en = open(folder + "/d_th(pr)-en.txt", "w")
    f_dpr_en = open(folder + "/d_th(dot+pr)-en.txt", "w")
    f_en_th = open(folder + "/d_en-th.txt", "w")

    # load the active sheet (there is only one) into the variable ws
    ws = wb2.active

    # speedup the processing with reset_dimensions
    ws.reset_dimensions()

    try:
        # dict-variable
        d = {}
        d_pr = {}
        d_th = {}
        # rows processed
        j = -1
        # rows saved
        count = 0
        for row in ws.values:
            j = j + 1
            li = []
            if j == 0 or j == 1:
                row = regex.sub(r'\\n.*?,', ',', str(row)+"',")
                print(row)
                continue
            # print(row[0])
            # breakpoint()
            # if there is no english definition: ignore row
            #  -> row 5 is english => row[4]
            # if there is no thai: ignore row 
            #  -> row 4 is thai => row[3]
            # lao is 25(lao) and 26(lao-roman)
            if 4 >= len(row) or str(row[4]) == "None" or str(row[3]) == "None":
                continue
            s = ""
            # for file in r:
            for i in range(0, columns - 1):
                if i >= len(row) or str(row[i]) == "None":
                    li.append("")
                    if i == 2:
                        li.append("")

                else:
                    # breakpoint
                    s = str(row[i])
                    if i == 2:  # i=2 is the 3rd column
                        st = format_tones(s.lower(), paiboon)
                        st2 = format_pron_search(st, paiboon)
                        st = spaces_workaround_dictbox(st)
                        li.append(st)
                        li.append(st2)

                        # breakpoint()
                        if j == 5000090:
                            print(st + " " + st2)
                            breakpoint()
                    else:
                        li.append(format_def(s))

                # i = i+1
            # print(l)
            count = count + 1
            # breakpoint()

            # for troubleshooting: save every column in separate file
            if debug:
                i = 0
                for file in r:
                    file.write(li[i] + "\n")
                    i = i + 1

            # word column for thai-eng
            w_th = li[4]

            # word column for thai(pron)-eng
            w_pr = ""
            if len(li[3]) > 0:
                w_pr = li[3] + " - " + li[4]

            # breakpoint()
            # definition (same for both)
            if len(li[11]) > 0:
                tmp = li[11].split(";")
                # if len(tmp)>1: breakpoint()
                t_cl = ""
                for t in tmp:
                    #  t_cl+= regex.sub(r'\s*(.*?)\s*\(([ก-๛]+).*\s*', '\\2 (\\1)', t) + ', '
                    t_cl += format_classifier(t, paiboon) + ", "
                    # breakpoint()

                w_clf = (
                    ' <clf style="font-size:0.8em">classifier: ' + t_cl[:-2] + "</clf>"
                )
            else:
                w_clf = ""
            tmp_pron = ""
            if len(li[2]) > 0:
                if paiboon:
                    tmp_pron = (
                        '<pron style="color:brown"> ['
                        + format_final_pron(li[2], paiboon)
                        + "] </pron>"
                    )
                else:
                    tmp_pron = "[" + format_final_pron(li[2], paiboon) + "]"
            w_def = (
                "<thai><b>"
                + li[4]
                + "</b></thai> "
                + tmp_pron
                + ' <type style="color:green">'
                + li[7].lower()
                + " "
                + li[8]
                + "</type> "
                + w_clf
                + "<br><def>"
                + li[5]
                + "</def><br>"
            )
            if len(li[12]) > 0:
                tmp = li[12].split(";")
                t_syn = ""
                for t in tmp:
                    t_syn += format_classifier(t, paiboon) + ", "
                # if len(tmp)>1: breakpoint()

                w_def += '<syn>syn: ' + t_syn[:-2] + "</syn><br>"
            if len(li[9]) > 0:
                w_def += (
                    '<science style="font-size:0.8em">scient: ' + li[9] + "</science><br>"
                )
            if len(li[14]) > 0:
                w_def += '<note style="font-size:0.8em">note: ' + li[14] + "</note><br>"
            level = l1 = l2 = l3 = ""

            if len(li[13]) > 0:
                l1 = "Level: " + li[13] + " "
            if len(li[13]) > 0 and len(li[10]) > 0:
                l2 = " - "
            if len(li[10]) > 0:
                l3 = "Category: " + li[10].lower()
            if len(l1) + len(l2) + len(l3) > 0:
                level = '<level style="font-size:0.7em">' + l1 + l2 + l3 + "</level>"
            w_def += level

            sort = "  "
            # small hack: each entry has 2 letters at the beginning for sorting by usage (easyer words first)
            if len(li[13]) > 1:
                sort = str(li[13])[0:2]
            if len(li[13]) == 1:
                sort = str(li[13]) + " "

            # f_th_en.write(w_th+'\t'+w_def+'\n')
            # f_pr_en.write(w_pr+'\t'+w_def+'\n')
            if d_th.get(w_th) is None:
                d_th[w_th] = [sort + w_def]
            else:
                d_th[w_th].append(sort + w_def)

            # if(w_th=="เสมือนจริง"): breakpoint()

            if len(w_pr) > 0:
                if d_pr.get(w_pr) is None:
                    d_pr[w_pr] = [sort + w_def]
                else:
                    d_pr[w_pr].append(sort + w_def)

            # en-th
            tmp = li[5].split(";")
            # if(len(li[2])>1): pron = '['+format_final_pron(li[2],paiboon)+ ']'
            w_en = (
                sort
                + "<thai><b>"
                + li[4]
                + "</b></thai> "
                + tmp_pron
                + "<def>"
                + li[5]
                + '</def> <level style="font-size:0.7em">Level: '
                + li[13]
                + "</level>"
            )
            for wo in tmp:
                w_w = wo.strip()
                w_t = li[7].lower()
                # breakpoint()
                if d.get(w_w) is None:
                    d[w_w] = {w_t: [w_en]}
                    # d[w_w][w_t].append(w_en)
                else:
                    if d[w_w].get(w_t) is None:
                        d[w_w][w_t] = [w_en]
                    else:
                        d[w_w][w_t].append(w_en)

            if j == 5:  # row 1&2 dont reach
                print("\nprocessed nr of rows: ", flush=True)
            if j % 1000 == 0:
                # print(d)
                # breakpoint()
                # break
                for file in r:
                    file.flush()
                print(str(j), end=" ", flush=True)

            if j == 1000:
                if debug_test_1000_rows:
                    print('only testing with 1000 rows')
                    break


        print("\nsum entries: " + str(count))

        for e_w, e_t in d_th.items():
            #  breakpoint()
            e_t.sort()
            for n_def in e_t:
                # print(str(x_def), str(n_type), str(e_w), sep=' | ' ) s_def=''
                f_th_en.write(e_w + "\t" + n_def[2:] + "\n")

        for e_w, e_t in d_pr.items():
            #  breakpoint()
            e_t.sort()
            for n_def in e_t:
                # print(str(x_def), str(n_type), str(e_w), sep=' | ' ) s_def=''
                f_pr_en.write(e_w + "\t" + n_def[2:] + "\n")
                f_dpr_en.write("." + e_w + "\t" + n_def[2:] + "\n")

        #      f_pr_en.write(w_pr+'\t'+w_def+'\n')

        s_type = ""
        s_def = ""
        for e_w, e_t in d.items():
            #  breakpoint()
            dict(sorted(e_t.items()))
            for n_type, n_def in e_t.items():
                # print(str(x_def), str(n_type), str(e_w), sep=' | ' ) s_def=''
                n_def.sort()
                for x_def in n_def:
                    # print(str(x_def), str(n_type), str(e_w), sep=' | ' )
                    s_def += x_def[2:] + "<br>"
                    # breakpoint()
                s_type += '<type style="color:green">' + n_type + "</type><br>" + s_def
                s_def = ""
            s_word = e_w + "\t<b>" + e_w + "</b><br> " + s_type
            s_type = ""

            # breakpoint()
            f_en_th.write(s_word + "\n")

    # test----
    # s_type=''
    # s_def=''
    # i=0
    # for e_w,e_t in d.items():
    # #  breakpoint()
    #   i+=1
    #   if (i>6): break
    #   for n_type,n_def in e_t.items():
    #    #print(str(x_def), str(n_type), str(e_w), sep=' | ' ) s_def=''
    #     for x_def in n_def:
    #       #print(str(x_def), str(n_type), str(e_w), sep=' | ' )
    #       s_def+='---'+x_def+'\n'
    #       #breakpoint()
    #     s_type+='--'+n_type+'\n'+s_def
    #     s_def=''
    #   s_word='-'+e_w+'\n'+s_type
    #   s_type=''
    #   #breakpoint()
    #   print(s_word)
    # test----end

    except Exception:
        pdb.post_mortem()
        raise

    finally:

        for file in r:
            file.close()


# default debug options
debug_test_1000_rows = False

#quick change
# debug_test_1000_rows = True

if __name__ == "__main__":
    import_excel_file(
        "vol_mundo_01.06.2023.xlsx",
        to_folder="vol_mundo",
        columns=30,
        paiboon=True,
        debug=False,
    )

    #auto convert to ifo file format and copy to dictboxdata
    #set the title and description and outputpath in here
    make_dict_files()
