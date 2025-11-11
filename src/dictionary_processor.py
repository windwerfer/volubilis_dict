"""Dictionary processing logic for Excel to text conversion."""

import hashlib
import logging
import pickle
import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    load_workbook = None

from .config import Config, DictionaryConfig
from .file_handler import FileHandler
from .text_formatter import TextFormatter


logger = logging.getLogger(__name__)


class DictionaryProcessor:
    """Processes Excel dictionary files into various output formats."""

    def __init__(self, config: Config):
        self.config = config.dictionary
        self.formatter = TextFormatter(self.config.patterns)
        self.file_handler = FileHandler()

        # Ensure cache file is in output directory
        if not self.config.cache_file.is_absolute():
            self.config.cache_file = self.config.output_folder / self.config.cache_file

    def process_excel_file(self) -> None:
        """Main method to process the Excel file."""
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not available - using mock data for demonstration")
            # Use mock processing for demonstration
            self._process_mock_data()
            return

        logger.info(f"Processing Excel file: {self.config.excel_file}")

        # Check cache first if enabled
        if self.config.use_cache and not self.config.force_refresh_cache:
            cached_data = self._load_from_cache()
            if cached_data is not None:
                logger.info("Loaded data from cache")
                self._write_cached_data_to_files(cached_data)
                return

        # Ensure output directory exists
        self.file_handler.ensure_directory(self.config.output_folder)

        # Load workbook
        wb = load_workbook(self.config.excel_file, read_only=True)
        logger.info(f"Sheet names: {wb.sheetnames}")

        # Process the active sheet
        ws = wb.active
        ws.reset_dimensions()

        # Log column mapping from header row (second row)
        rows = list(ws.values)
        if len(rows) > 1:
            header_row = rows[1]  # Second row (index 1)
            mapping_lines = []
            for i, col in enumerate(header_row):
                usage = "unused"
                if i == 0: usage = "thai_romanized"
                elif i == 1: usage = "easythai"
                elif i == 2: usage = "thaiphon (pronunciation)"
                elif i == 3: usage = "thai (word)"
                elif i == 4: usage = "english (definition)"
                elif i == 7: usage = "type_word"
                elif i == 8: usage = "scient"
                elif i == 9: usage = "dom"
                elif i == 10: usage = "classif"
                elif i == 11: usage = "syn"
                elif i == 12: usage = "level"
                elif i == 13: usage = "note"
                col_clean = str(col).replace('\n', ' ')
                mapping_lines.append(f"{i}: {col_clean} -> {usage}")
            logger.info("Column mapping:\n" + "\n".join(mapping_lines))

        # Initialize data structures
        th_en_data = defaultdict(list)  # Thai to English
        th_pron_en_data = defaultdict(list)  # Thai with pronunciation to English
        th_pron_merge_en_data = defaultdict(list)  # Merged pronunciation to English
        en_th_data = defaultdict(lambda: defaultdict(list))  # English to Thai

        # Open output files
        output_files = self._open_output_files()

        try:
            row_count = 0
            processed_count = 0

            for row in ws.values:
                row_count += 1

                # Skip header rows
                if row_count <= 2:
                    continue

                # Process the row
                processed = self._process_row(row, th_en_data, th_pron_en_data,
                                            th_pron_merge_en_data, en_th_data)
                if processed:
                    processed_count += 1

                # Progress logging
                if row_count == 6:
                    logger.info("Processing rows...")
                if row_count % 1000 == 0:
                    logger.info(f"Processed {row_count} rows")

                # Debug limit
                if self.config.debug_test_1000_rows and row_count >= 1000:
                    logger.info("Debug mode: stopping at 1000 rows")
                    break

            logger.info(f"Total entries processed: {processed_count}")

            # Write the processed data to files
            self._write_output_files(output_files, th_en_data, th_pron_en_data,
                                   th_pron_merge_en_data, en_th_data)

        finally:
            # Close all files
            for f in output_files.values():
                f.close()

        # Save to cache if enabled
        if self.config.use_cache:
            # Convert defaultdict structures to regular dicts for pickling
            cache_data = {
                'th_en': self._convert_defaultdict_to_dict(th_en_data),
                'th_pron_en': self._convert_defaultdict_to_dict(th_pron_en_data),
                'en_th': self._convert_defaultdict_to_dict(en_th_data)
            }
            self._save_to_cache(cache_data)

    def _process_mock_data(self) -> None:
        """Process mock data for demonstration when openpyxl is not available."""
        logger.info("Processing mock dictionary data for demonstration")

        # Check cache first
        if self.config.use_cache and not self.config.force_refresh_cache:
            cached_data = self._load_from_cache()
            if cached_data is not None:
                logger.info("Loaded mock data from cache")
                self._write_cached_data_to_files(cached_data)
                return

        # Ensure output directory exists
        self.file_handler.ensure_directory(self.config.output_folder)

        # Create mock data as tuples (row[3]=Thai, row[5]=English are required)
        # Include pronunciation data (row[2]) to enable English→Thai processing
        mock_data = [
            ('', '', 'sà-wàt-dii', 'สวัสดี', 'sawadee', 'hello', '', 'greeting', 'common', '', '', '', '', 'A1', ''),
            ('', '', 'khòp-khùn', 'ขอบคุณ', 'khopkhun', 'thank you', '', 'expression', 'common', '', '', '', '', 'A1', ''),
            ('', '', 'mɛɛw', 'แมว', 'maew', 'cat', '', 'noun', 'common', '', 'animal', 'ตัว', '', 'A1', ''),
            ('', '', 'sù-nák', 'สุนัข', 'sunak', 'dog', '', 'noun', 'common', '', 'animal', 'ตัว', '', 'A1', ''),
            ('', '', 'bâan', 'บ้าน', 'ban', 'house', '', 'noun', 'common', '', 'building', 'หลัง', '', 'A1', ''),
        ]

        # Initialize data structures
        th_en_data = defaultdict(list)
        th_pron_en_data = defaultdict(list)
        th_pron_merge_en_data = defaultdict(list)
        en_th_data = defaultdict(lambda: defaultdict(list))

        # Process mock rows
        for row in mock_data:
            self._process_row(tuple(row), th_en_data, th_pron_en_data,
                             th_pron_merge_en_data, en_th_data)

        # Open output files and write data
        output_files = self._open_output_files()

        try:
            self._write_output_files(output_files, th_en_data, th_pron_en_data,
                                   th_pron_merge_en_data, en_th_data)
        finally:
            for f in output_files.values():
                f.close()

        # Save to cache
        if self.config.use_cache:
            cache_data = {
                'th_en': dict(th_en_data),
                'th_pron_en': dict(th_pron_en_data),
                'en_th': {k: dict(v) for k, v in en_th_data.items()}
            }
            self._save_to_cache(cache_data)

        logger.info("Mock processing completed - 5 sample entries created")

    def _open_output_files(self) -> Dict[str, any]:
        """Open all output files."""
        base_path = self.config.output_folder

        files = {
            'th_en': open(base_path / "volubilis_th-en.txt", "w", encoding='utf-8'),
            'th_pron_en': open(base_path / "volubilis_th-pr-en.txt", "w", encoding='utf-8'),
            'en_th': open(base_path / "volubilis_en-th.txt", "w", encoding='utf-8'),
        }
        if self.config.th_pron_merge:
            files['th_pron_merge_en'] = open(base_path / "volubilis_th-pr-merge-en.txt", "w", encoding='utf-8')
        return files

    def _process_row(
        self,
        row: Tuple,
        th_en_data: Dict,
        th_pron_en_data: Dict,
        th_pron_merge_en_data: Dict,
        en_th_data: Dict
    ) -> bool:
        """Process a single row from the Excel file."""
        # Validate required columns
        if len(row) < 5 or not row[4] or not row[3]:  # English and Thai required
            return False

        # Extract column values
        thai_romanized = self.formatter.clean_text(row[0] if len(row) > 0 else "")
        easythai = self.formatter.clean_text(row[1] if len(row) > 1 else "")
        thaiphon = self.formatter.clean_text(row[2] if len(row) > 2 else "")
        thai = self.formatter.clean_text(row[3])
        english = self.formatter.clean_text(row[4])

        # Split synonyms
        thai_synonyms = [s.strip() for s in re.split(r'[;=]', thai) if s.strip()]
        english_synonyms = [s.strip() for s in re.split(r'[;=]', english) if s.strip()]

        # Extract additional synonyms from SYN column (Thai words in parentheses)
        syn = self.formatter.clean_text(row[11] if len(row) > 11 else "")
        if syn:
            bracketed_matches = re.findall(r'\((.*?)\)', syn)
            for match in bracketed_matches:
                cleaned_match = match.strip()
                if cleaned_match:
                    # Check for Thai characters
                    if re.search(r'[\u0E00-\u0E7F]', cleaned_match):
                        # Ignore if contains spaces or invalid chars
                        if ' ' not in cleaned_match and not re.search(r'[^\u0E00-\u0E7F]', cleaned_match):
                            if cleaned_match not in thai_synonyms:
                                thai_synonyms.append(cleaned_match)

        # Use joined synonyms for word keys
        thai_word = '|'.join(thai_synonyms) if thai_synonyms else thai
        english_word = '|'.join(english_synonyms) if english_synonyms else english

        # Use first synonym for display in definitions
        thai_display = thai_synonyms[0] if thai_synonyms else thai

        # Additional columns
        type_word = self.formatter.clean_text(row[7] if len(row) > 7 else "")
        usage = self.formatter.clean_text(row[8] if len(row) > 8 else "")
        scient = self.formatter.clean_text(row[9] if len(row) > 9 else "")
        dom = self.formatter.clean_text(row[10] if len(row) > 10 else "")
        classif = self.formatter.clean_text(row[10] if len(row) > 10 else "")
        syn = self.formatter.clean_text(row[11] if len(row) > 11 else "")
        level = self.formatter.clean_text(row[13] if len(row) > 13 else "")
        note = self.formatter.clean_text(row[14] if len(row) > 14 else "")

        # Format pronunciation
        pron_formatted = self.formatter.format_tones(thaiphon.lower(), self.config.paiboon)
        pron_search = self.formatter.format_pronunciation_search(pron_formatted, self.config.paiboon)

        # Create pronunciation headword
        if self.config.th_pron_incl_translation_in_headword:
            eng_summary = english_word.replace('|', ', ')[:50]
            pron_headword = f"{pron_search} - {thai_word} ({eng_summary})"
        else:
            pron_headword = f"{pron_search} - {thai_word}"

        # Truncate headword if too long
        pron_headword = pron_headword[:self.config.th_pron_max_headword_length]

        # Format definition
        definition = self._format_definition(
            thai_display, pron_formatted, type_word, usage, classif, syn, scient, note, level, english_word, dom
        )

        # Collect for pron merge
        if self.config.th_pron_merge:
            for i, thai_syn in enumerate(thai_synonyms):
                eng_syn = english_synonyms[i] if i < len(english_synonyms) else ""
                base_pron = pron_search
                if base_pron:
                    th_pron_merge_en_data[base_pron].append((thai_syn, eng_syn, level, definition))

        # Add sorting prefix for Thai-English
        sort_prefix = self._get_sort_prefix(level)

        # Thai to English entries
        th_en_data[thai_word].append(sort_prefix + definition)

        # Thai with pronunciation to English
        if pron_headword:
            th_pron_en_data[pron_headword].append(sort_prefix + definition)

            # English to Thai entries
            self._add_english_to_thai_entries(english_word, definition, type_word, en_th_data)

        return True

    def _convert_defaultdict_to_dict(self, d):
        """Recursively convert defaultdict structures to regular dicts for pickling."""
        if isinstance(d, dict):
            return {k: self._convert_defaultdict_to_dict(v) for k, v in d.items()}
        return d

    def _generate_cache_key(self) -> str:
        """Generate a cache key based on file modification time and configuration."""
        # Get file modification time
        if self.config.excel_file.exists():
            mtime = self.config.excel_file.stat().st_mtime
        else:
            mtime = 0

        # Create a hash of relevant configuration
        config_str = f"{self.config.columns}_{self.config.paiboon}_{self.config.debug_test_1000_rows}_{mtime}"
        return hashlib.md5(config_str.encode()).hexdigest()

    def _save_to_cache(self, data: Dict[str, Any]) -> None:
        """Save processed data to cache file."""
        try:
            # Ensure cache directory exists
            self.config.cache_file.parent.mkdir(parents=True, exist_ok=True)

            cache_data = {
                'cache_key': self._generate_cache_key(),
                'data': data
            }
            with open(self.config.cache_file, 'wb') as f:
                pickle.dump(cache_data, f)
            logger.info(f"Saved data to cache: {self.config.cache_file}")
        except Exception as e:
            logger.warning(f"Failed to save cache: {e}")

    def _load_from_cache(self) -> Optional[Dict[str, Any]]:
        """Load processed data from cache if valid."""
        try:
            if not self.config.cache_file.exists():
                return None

            with open(self.config.cache_file, 'rb') as f:
                cache_data = pickle.load(f)

            # Check if cache is still valid
            if cache_data.get('cache_key') == self._generate_cache_key():
                logger.info("Cache is valid, loading data")
                return cache_data['data']
            else:
                logger.info("Cache is outdated, will reprocess")
                return None

        except Exception as e:
            logger.warning(f"Failed to load cache: {e}")
            return None

    def _write_cached_data_to_files(self, data: Dict[str, Any]) -> None:
        """Write cached data to output files."""
        th_en_data = data['th_en']
        th_pron_en_data = data['th_pron_en']
        th_dot_pron_en_data = data['th_dot_pron_en']
        en_th_data = data['en_th']
        th_pron_merge_en_data = defaultdict(list)  # New feature, not in cache

        # Open output files
        output_files = self._open_output_files()

        try:
            # Write the processed data to files
            self._write_output_files(output_files, th_en_data, th_pron_en_data,
                                   th_pron_merge_en_data, en_th_data)
        finally:
            # Close all files
            for f in output_files.values():
                f.close()

    def _format_definition(
        self,
        thai: str,
        pron_formatted: str,
        type_word: str,
        usage: str,
        classif: str,
        syn: str,
        scient: str,
        note: str,
        level: str,
        english: str,
        dom: str = ""
    ) -> str:
        """Format a complete definition string with standard HTML and CSS classes."""
        definition = f'<span class="thai"><strong>{thai}</strong></span> '

        # Add pronunciation
        if pron_formatted:
            if self.config.paiboon:
                definition += f'<span class="pron">[{self.formatter.format_final_pronunciation(pron_formatted, self.config.paiboon)}]</span> '
            else:
                definition += f'<span class="pron">[{self.formatter.format_final_pronunciation(pron_formatted, self.config.paiboon)}]</span> '

        # Add type and usage
        type_usage = f"{type_word.lower()} {usage}".strip()
        if type_usage:
            definition += f'<span class="type">{type_usage}</span> '

        # Add classifier
        if classif:
            classifiers = self.formatter.split_and_format_classifiers(classif, self.config.paiboon)
            if classifiers.strip():
                definition += f'<span class="clf">classifier: {classifiers}</span> '

        # Add definition
        english_formatted = english.replace("|", ", ").replace(";", ", ")
        definition += f'<br><span class="def">{english_formatted}</span><br>'

        # Add synonyms
        if syn:
            synonyms = self.formatter.split_and_format_synonyms(syn, self.config.paiboon)
            if synonyms.strip():
                definition += f'<span class="syn">syn: {synonyms}</span><br>'

        # Add scientific name
        if scient and scient.strip():
            definition += f'<span class="science">scient: {scient.replace("|", ", ")}</span><br>'

        # Add note
        if note and note.strip():
            definition += f'<span class="note">note: {note}</span><br>'

        # Add level
        level_info = self._format_level_info(level, dom)
        if level_info:
            definition += level_info

        return definition

    def _get_sort_prefix(self, level: str) -> str:
        """Get sorting prefix based on level."""
        if len(level) > 1:
            return level[:2]
        elif len(level) == 1:
            return level + " "
        return "  "

    def _format_level_info(self, level: str, dom: str) -> str:
        """Format level and domain information with standard HTML."""
        parts = []
        if level:
            parts.append(f"Level: {level}")
        if level and dom:
            parts.append(" - ")
        if dom:
            parts.append(f"Category: {dom.lower()}")

        if parts:
            return f'<span class="level">{"".join(parts)}</span>'
        return ""

    def _add_english_to_thai_entries(
        self,
        english: str,
        definition: str,
        type_word: str,
        en_th_data: Dict
    ) -> None:
        """Add entries to English to Thai data structure."""
        # For en-th, remove science span
        definition = re.sub(r'<span class="science">scient: (.*?)</span>', '', definition)

        # Set description to full ENG field for complete meaning context, replace | with ,
        definition = re.sub(r'<span class="description">.*?</span>', f'<span class="description">{english.replace("|", ", ")}</span>', definition)

        english_terms = [term.strip() for term in english.split("|") if term.strip()]

        for term in english_terms:
            en_th_data[term][type_word].append(definition)

    def _write_output_files(self, files, th_en_data, th_pron_en_data, th_pron_merge_en_data, en_th_data):
        """Write all processed data to output files."""
        # Thai to English
        for thai_word, definitions in th_en_data.items():
            definitions.sort()
            for definition in definitions:
                files['th_en'].write(f"{thai_word}\t{definition[2:]}\n")

        # Thai pronunciation to English
        if self.config.th_pron:
            for pron_word, definitions in th_pron_en_data.items():
                definitions.sort()
                key = self.config.th_pron_prefix + pron_word if self.config.th_pron_prefix else pron_word
                for definition in definitions:
                    if self.config.th_pron_incl_translation_in_headword:
                        # Extract English from definition or use full
                        # For simplicity, use the definition as is, but perhaps modify pron_headword to include eng
                        # Wait, currently pron_entry is pron - thai, definition is the full def
                        # To include eng, perhaps change pron_entry to pron - thai (eng)
                        # But since eng is in definition, maybe keep as is, or adjust
                        # For now, since definition includes eng, just write as is
                        files['th_pron_en'].write(f"{key}\t{definition[2:]}\n")
                    else:
                        # Remove eng from definition? But complicated.
                        # For now, assume if not incl, just thai
                        # But to keep simple, always include for now
                        files['th_pron_en'].write(f"{key}\t{definition[2:]}\n")

        # Thai pronunciation merge to English
        if self.config.th_pron_merge and 'th_pron_merge_en' in files:
            for base_pron, items in th_pron_merge_en_data.items():
                # Sort items by thai, then level
                sorted_items = self.formatter.sort_thai_words_by_tone_and_level(items, self._get_sort_prefix)
                key = self.config.th_pron_merge_prefix + base_pron if self.config.th_pron_merge_prefix else base_pron
                key = key[:self.config.th_pron_merge_max_headword_length]
                thai_list = []
                for thai, eng, _, _ in sorted_items:
                    if self.config.th_pron_merge_incl_translation_in_headword and eng:
                        thai_list.append(f"{thai} ({eng})")
                    else:
                        thai_list.append(thai)
                headword_part = ", ".join(thai_list)
                key = self.config.th_pron_merge_prefix + base_pron + " - " + headword_part
                key = key[:self.config.th_pron_merge_max_headword_length]
                # Merge all definitions with <br><br> separator
                definitions = [item[3] for item in sorted_items if item[3]]
                value = "<br><br>".join(definitions) if definitions else ""
                files['th_pron_merge_en'].write(f"{key}\t{value}\n")

        # English to Thai
        import re
        for english_word, type_groups in en_th_data.items():
            # Sort types
            sorted_types = sorted(type_groups.items())

            type_entries = []
            for word_type, definitions in sorted_types:
                definitions.sort()
                def_text = "<br>".join(definition for definition in definitions)
                # Remove the English def span for cleaner en-th
                def_text = re.sub(r'<br><span class="def">.*?</span><br>', '', def_text)
                if word_type.strip():
                    type_entries.append(f'<span class="type">{word_type}</span><br>{def_text}')
                else:
                    type_entries.append(def_text)

            word_definition = f'<span class="english"><strong>{english_word}</strong></span> <br>' + "<br>".join(type_entries)
            files['en_th'].write(f"{english_word}\t{word_definition}\n")